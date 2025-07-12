from fastapi import FastAPI, APIRouter, Query, HTTPException, Response
import fdb
import os
from dotenv import load_dotenv
from datetime import datetime
from fastapi.responses import JSONResponse
from io import StringIO
import csv

# Carregar variáveis do .env
load_dotenv()

app = FastAPI()
router = APIRouter()

def get_firebird_conn():
    return fdb.connect(
        host=os.getenv("FIREBIRD_HOST"),
        database=os.getenv("FIREBIRD_DATABASE"),
        user=os.getenv("FIREBIRD_USER"),
        password=os.getenv("FIREBIRD_PASSWORD"),
        charset=os.getenv("FIREBIRD_CHARSET"),
    )

def parse_date(date_str, param_name):
    try:
        # Aceita tanto 'DD/MM/YYYY' quanto 'YYYY-MM-DD'
        if "/" in date_str:
            return datetime.strptime(date_str, "%d/%m/%Y").date()
        else:
            return datetime.strptime(date_str, "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(status_code=422, detail=f"Parâmetro de data inválido: {param_name} deve ser DD/MM/AAAA")

# --- ROTA DE CONSULTA (JSON PAGINADO) ---
@router.get("/consulta")
def consulta_cobranca(
    response: Response,
    pagamento_ini: str = Query(None),
    pagamento_fim: str = Query(None),
    vencimento_ini: str = Query(None),
    vencimento_fim: str = Query(None),
    nome: str = Query(None),
    pagina: int = Query(1, ge=1),
    limite: int = Query(100, ge=1, le=1000)
):
    filtros = []
    params = []
    # Filtros robustos (só aplica se tem valor e não vazio)
    if pagamento_ini and pagamento_ini.strip():
        dt_ini = parse_date(pagamento_ini, "pagamento_ini")
        filtros.append("CAST(RECEBER.pagamento AS DATE) >= ?")
        params.append(dt_ini)
    if pagamento_fim and pagamento_fim.strip():
        dt_fim = parse_date(pagamento_fim, "pagamento_fim")
        filtros.append("CAST(RECEBER.pagamento AS DATE) <= ?")
        params.append(dt_fim)
    if vencimento_ini and vencimento_ini.strip():
        vdt_ini = parse_date(vencimento_ini, "vencimento_ini")
        filtros.append("CAST(RECEBER.vencimento AS DATE) >= ?")
        params.append(vdt_ini)
    if vencimento_fim and vencimento_fim.strip():
        vdt_fim = parse_date(vencimento_fim, "vencimento_fim")
        filtros.append("CAST(RECEBER.vencimento AS DATE) <= ?")
        params.append(vdt_fim)
    if nome and nome.strip():
        filtros.append("UPPER(TRIM(t.razao)) LIKE ?")
        params.append(f"%{nome.strip().upper()}%")

    offset = (pagina - 1) * limite

    sql = f"""
    SELECT FIRST {limite} SKIP {offset}
        RECEBER.numero, 
        SUBSTRING(t.razao FROM 1 FOR 30) AS nome, 
        RECEBER.valor_nf, 
        RECEBER.emissao, 
        RECEBER.vencimento, 
        RECEBER.valor, 
        RECEBER.pagamento, 
        RECEBER.valor_pago, 
        RECEBER.referente_a, 
        c.dt_cancelado, 
        t.cgc,
        t.telefone, 
        t.fax, 
        t.fone_comercial, 
        t.endereco, 
        t.bairro, 
        t.cidade, 
        t.uf, 
        t.cep, 
        t.email, 
        t.obs, 
        c.nr_contrato,  
        c.nr_terreno,   
        RECEBER.conta, 
        pl.descricao AS plano_classif, 
        pl.descricao AS plano_descricao, 
        CASE 
            WHEN RECEBER.tipo = 1 THEN 'Manutenção' 
            WHEN RECEBER.tipo = 2 THEN 'Venda' 
            WHEN RECEBER.tipo = 3 THEN 'Jazigo' 
            WHEN RECEBER.tipo = 4 THEN 'Outros' 
            WHEN RECEBER.tipo = 5 THEN 'Produtos' 
            ELSE 'Outros' 
        END AS tipo_descricao 
    FROM RECEBER 
    JOIN TITULAR t ON t.Codigo = RECEBER.TITULAR 
    JOIN CONTRATO c ON c.nr_contrato = RECEBER.nf 
    JOIN PLANO pl ON pl.codigo = RECEBER.conta 
    """
    if filtros:
        sql += "\nWHERE " + " AND ".join(filtros)

    sql_count = """
    SELECT COUNT(*)
    FROM RECEBER 
    JOIN TITULAR t ON t.Codigo = RECEBER.TITULAR 
    JOIN CONTRATO c ON c.nr_contrato = RECEBER.nf 
    JOIN PLANO pl ON pl.codigo = RECEBER.conta 
    """
    if filtros:
        sql_count += "\nWHERE " + " AND ".join(filtros)

    try:
        conn = get_firebird_conn()
        cur = conn.cursor()
        try:
            cur.execute(sql_count, params)
            total_registros = cur.fetchone()[0]
            cur.execute(sql, params)
            columns = [desc[0].lower() for desc in cur.description]
            rows = cur.fetchall()
        finally:
            try: cur.close()
            except: pass
            try: conn.close()
            except: pass

        result = []
        for row in rows:
            d = {}
            for col, val in zip(columns, row):
                d[col] = val
            result.append(d)

        return {
            "status": "ok",
            "pagina": pagina,
            "limite": limite,
            "total_registros": total_registros,
            "dados": result
        }
    except Exception as e:
        response.status_code = 500
        return {"status": "erro", "mensagem": f"Erro ao executar consulta: {str(e)}"}

# --- ROTA DE EXPORTAÇÃO CSV ---

def tratar_valor_csv(col, v):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return ""
    # Datas
    if col in {"emissao", "vencimento", "pagamento", "dt_cancelado"} and isinstance(v, (datetime, )):
        return v.strftime('%d/%m/%Y')
    # Moeda
    if col in {"valor_nf", "valor", "valor_pago"}:
        try:
            valor = float(v)
            valor_str = f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            return f'R$ {valor_str}'
        except:
            return v
    # CGC/CEP/NÚMEROS: SEM apóstrofo, como texto puro
    if col in {"cgc", "cep", "nr_terreno", "conta", "nr_contrato"}:
        return str(v).strip()
    # Texto normal
    return str(v)

from fastapi import APIRouter, Query, Response
from fastapi.responses import JSONResponse
from datetime import datetime
from io import BytesIO
from urllib.parse import quote
import openpyxl

# ... (demais imports, get_firebird_conn, parse_date, etc.)

@router.get("/consulta_xlsx")
def exporta_xlsx_cobranca(
    pagamento_ini: str = Query(None),
    pagamento_fim: str = Query(None),
    vencimento_ini: str = Query(None),
    vencimento_fim: str = Query(None),
    nome: str = Query(None)
):
    filtros = []
    params = []
    # Filtros robustos
    if pagamento_ini and pagamento_ini.strip():
        dt_ini = parse_date(pagamento_ini, "pagamento_ini")
        filtros.append("CAST(RECEBER.pagamento AS DATE) >= ?")
        params.append(dt_ini)
    if pagamento_fim and pagamento_fim.strip():
        dt_fim = parse_date(pagamento_fim, "pagamento_fim")
        filtros.append("CAST(RECEBER.pagamento AS DATE) <= ?")
        params.append(dt_fim)
    if vencimento_ini and vencimento_ini.strip():
        vdt_ini = parse_date(vencimento_ini, "vencimento_ini")
        filtros.append("CAST(RECEBER.vencimento AS DATE) >= ?")
        params.append(vdt_ini)
    if vencimento_fim and vencimento_fim.strip():
        vdt_fim = parse_date(vencimento_fim, "vencimento_fim")
        filtros.append("CAST(RECEBER.vencimento AS DATE) <= ?")
        params.append(vdt_fim)
    if nome and nome.strip():
        filtros.append("UPPER(TRIM(t.razao)) LIKE ?")
        params.append(f"%{nome.strip().upper()}%")

    sql = """
    SELECT 
        RECEBER.numero, 
        SUBSTRING(t.razao FROM 1 FOR 30) AS nome, 
        RECEBER.valor_nf, 
        RECEBER.emissao, 
        RECEBER.vencimento, 
        RECEBER.valor, 
        RECEBER.pagamento, 
        RECEBER.valor_pago, 
        RECEBER.referente_a, 
        c.dt_cancelado, 
        t.cgc,
        t.telefone, 
        t.fax, 
        t.fone_comercial, 
        t.endereco, 
        t.bairro, 
        t.cidade, 
        t.uf, 
        t.cep, 
        t.email, 
        t.obs, 
        c.nr_contrato,  
        c.nr_terreno,   
        RECEBER.conta, 
        pl.descricao AS plano_classif, 
        pl.descricao AS plano_descricao, 
        CASE 
            WHEN RECEBER.tipo = 1 THEN 'Manutenção' 
            WHEN RECEBER.tipo = 2 THEN 'Venda' 
            WHEN RECEBER.tipo = 3 THEN 'Jazigo' 
            WHEN RECEBER.tipo = 4 THEN 'Outros' 
            WHEN RECEBER.tipo = 5 THEN 'Produtos' 
            ELSE 'Outros' 
        END AS tipo_descricao 
    FROM RECEBER 
    JOIN TITULAR t ON t.Codigo = RECEBER.TITULAR 
    JOIN CONTRATO c ON c.nr_contrato = RECEBER.nf 
    JOIN PLANO pl ON pl.codigo = RECEBER.conta 
    """
    if filtros:
        sql += "\nWHERE " + " AND ".join(filtros)

    try:
        conn = get_firebird_conn()
        cur = conn.cursor()
        cur.execute(sql, params)
        columns = [desc[0].upper() for desc in cur.description]  # Cabeçalho em maiúsculo
        rows = cur.fetchall()
        cur.close()
        conn.close()

        # Índices das colunas que devem ser tratadas
        moeda_cols = {"VALOR_NF", "VALOR", "VALOR_PAGO"}
        texto_cols = {"CGC", "CEP", "NR_TERRENO", "CONTA", "NR_CONTRATO"}
        data_cols = {"EMISSAO", "VENCIMENTO", "PAGAMENTO", "DT_CANCELADO"}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "COBRANCA"

        ws.append(columns)
        for row in rows:
            formatted = []
            for colname, val in zip(columns, row):
                if colname in data_cols and isinstance(val, (datetime, )):
                    formatted.append(val.strftime('%d/%m/%Y'))
                elif colname in moeda_cols and val is not None:
                    try:
                        formatted.append(float(val))
                    except:
                        formatted.append(val)
                elif colname in texto_cols and val is not None:
                    formatted.append(str(val).strip())
                else:
                    formatted.append(val)
            ws.append(formatted)

        # Aplica formatação de moeda nas colunas de valor
        for idx, colname in enumerate(columns, start=1):
            if colname in moeda_cols:
                for cell in ws[openpyxl.utils.get_column_letter(idx)][1:]:  # Pula cabeçalho
                    cell.number_format = u'R$ #,##0.00'

        # Geração dinâmica do nome do arquivo conforme os filtros
        nome_base = "cobranca_bonfim"
        partes_nome = []
        if pagamento_ini and pagamento_fim:
            partes_nome.append(f"pgto_{pagamento_ini.replace('/','-')}_a_{pagamento_fim.replace('/','-')}")
        if vencimento_ini and vencimento_fim:
            partes_nome.append(f"vcto_{vencimento_ini.replace('/','-')}_a_{vencimento_fim.replace('/','-')}")
        if nome and nome.strip():
            partes_nome.append(f"nome_{nome.strip().replace(' ', '_')}")
        nome_arquivo = f"{nome_base}_{'_'.join(partes_nome) or 'completo'}.xlsx"
        nome_arquivo = quote(nome_arquivo)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{nome_arquivo}"
            }
        )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "erro", "mensagem": f"Erro ao exportar XLSX: {str(e)}"}
        )

from fastapi import Query

@router.get("/renegociacao")
def consulta_renegociacao(
    response: Response,
    pagina: int = Query(1, ge=1),
    limite: int = Query(100, ge=1, le=1000),
    cnpj: str = Query(None),
    contrato: str = Query(None),
):
    filtros = ["RECEBER.conta IN (306, 307)"]
    params = []
    if cnpj and cnpj.strip():
        filtros.append("t.cgc = ?")
        params.append(cnpj.strip())
    if contrato and contrato.strip():
        filtros.append("c.nr_contrato = ?")
        params.append(contrato.strip())

    offset = (pagina - 1) * limite

    sql = f"""
    SELECT FIRST {limite} SKIP {offset}
        RECEBER.numero, 
        SUBSTRING(t.razao FROM 1 FOR 30) AS nome, 
        RECEBER.valor_nf, 
        RECEBER.emissao, 
        RECEBER.vencimento, 
        RECEBER.valor, 
        RECEBER.pagamento, 
        RECEBER.valor_pago, 
        RECEBER.referente_a, 
        c.dt_cancelado, 
        t.cgc,
        t.telefone, 
        t.fax, 
        t.fone_comercial, 
        t.endereco, 
        t.bairro, 
        t.cidade, 
        t.uf, 
        t.cep, 
        t.email, 
        t.obs, 
        c.nr_contrato,  
        c.nr_terreno,   
        RECEBER.conta, 
        pl.descricao AS plano_classif, 
        pl.descricao AS plano_descricao, 
        CASE 
            WHEN RECEBER.tipo = 1 THEN 'Manutenção' 
            WHEN RECEBER.tipo = 2 THEN 'Venda' 
            WHEN RECEBER.tipo = 3 THEN 'Jazigo' 
            WHEN RECEBER.tipo = 4 THEN 'Outros' 
            WHEN RECEBER.tipo = 5 THEN 'Produtos' 
            ELSE 'Outros' 
        END AS tipo_descricao 
    FROM RECEBER 
    JOIN TITULAR t ON t.Codigo = RECEBER.TITULAR 
    JOIN CONTRATO c ON c.nr_contrato = RECEBER.nf 
    JOIN PLANO pl ON pl.codigo = RECEBER.conta 
    WHERE {" AND ".join(filtros)}
    ORDER BY RECEBER.vencimento
    """

    sql_count = f"""
    SELECT COUNT(*)
    FROM RECEBER 
    JOIN TITULAR t ON t.Codigo = RECEBER.TITULAR 
    JOIN CONTRATO c ON c.nr_contrato = RECEBER.nf 
    JOIN PLANO pl ON pl.codigo = RECEBER.conta 
    WHERE {" AND ".join(filtros)}
    """

    try:
        conn = get_firebird_conn()
        cur = conn.cursor()
        try:
            cur.execute(sql_count, params)
            total_registros = cur.fetchone()[0]
            cur.execute(sql, params)
            columns = [desc[0].lower() for desc in cur.description]
            rows = cur.fetchall()
        finally:
            try: cur.close()
            except: pass
            try: conn.close()
            except: pass

        result = []
        for row in rows:
            d = {}
            for col, val in zip(columns, row):
                d[col] = val
            result.append(d)

        return {
            "status": "ok",
            "pagina": pagina,
            "limite": limite,
            "total_registros": total_registros,
            "dados": result
        }
    except Exception as e:
        response.status_code = 500
        return {"status": "erro", "mensagem": f"Erro ao executar consulta de renegociação: {str(e)}"}

        from fastapi.responses import Response, JSONResponse
from io import BytesIO
import openpyxl

@router.get("/renegociacao_xlsx")
def exporta_xlsx_renegociacao(
    cnpj: str = Query(None),
    contrato: str = Query(None)
):
    filtros = ["RECEBER.conta IN (306, 307)"]
    params = []
    if cnpj and cnpj.strip():
        filtros.append("t.cgc = ?")
        params.append(cnpj.strip())
    if contrato and contrato.strip():
        filtros.append("c.nr_contrato = ?")
        params.append(contrato.strip())

    sql = f"""
    SELECT 
        RECEBER.numero, 
        SUBSTRING(t.razao FROM 1 FOR 30) AS nome, 
        RECEBER.valor_nf, 
        RECEBER.emissao, 
        RECEBER.vencimento, 
        RECEBER.valor, 
        RECEBER.pagamento, 
        RECEBER.valor_pago, 
        RECEBER.referente_a, 
        c.dt_cancelado, 
        t.cgc,
        t.telefone, 
        t.fax, 
        t.fone_comercial, 
        t.endereco, 
        t.bairro, 
        t.cidade, 
        t.uf, 
        t.cep, 
        t.email, 
        t.obs, 
        c.nr_contrato,  
        c.nr_terreno,   
        RECEBER.conta, 
        pl.descricao AS plano_classif, 
        pl.descricao AS plano_descricao, 
        CASE 
            WHEN RECEBER.tipo = 1 THEN 'Manutenção' 
            WHEN RECEBER.tipo = 2 THEN 'Venda' 
            WHEN RECEBER.tipo = 3 THEN 'Jazigo' 
            WHEN RECEBER.tipo = 4 THEN 'Outros' 
            WHEN RECEBER.tipo = 5 THEN 'Produtos' 
            ELSE 'Outros' 
        END AS tipo_descricao 
    FROM RECEBER 
    JOIN TITULAR t ON t.Codigo = RECEBER.TITULAR 
    JOIN CONTRATO c ON c.nr_contrato = RECEBER.nf 
    JOIN PLANO pl ON pl.codigo = RECEBER.conta 
    WHERE {" AND ".join(filtros)}
    ORDER BY RECEBER.vencimento
    """

    try:
        conn = get_firebird_conn()
        cur = conn.cursor()
        cur.execute(sql, params)
        columns = [desc[0].upper() for desc in cur.description]
        rows = cur.fetchall()
        cur.close()
        conn.close()

        # Criação do XLSX
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RENEGOCIACAO"  # <-- Nome fixo e seguro da aba

        ws.append(columns)
        for row in rows:
            ws.append(row)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        nome_arquivo = "parcelas_renegociacao.xlsx"

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{nome_arquivo}"
            }
        )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "erro", "mensagem": f"Erro ao exportar XLSX: {str(e)}"}
        )


app.include_router(router, prefix="/api/cobranca")